#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @FileName  :do_excel.py
# @Time      :2020/6/8 15:43
# @Author    :麻花

from openpyxl import load_workbook
from tools.read_config import ReadConfig
from tools.project_path import *
from tools.get_data import GetData
from tools.my_log import MyLog
from tools.do_regx import DoRegx

my_logger=MyLog()


class DoExcel():

    @classmethod
    def get_data(cls,filename):
        wb = load_workbook(filename)
        test_data = []
        mode = eval(ReadConfig.get_config(case_config_path,'MODE','mode'))

        NoRegTel = getattr(GetData,'NoRegTel')
        for key in mode :
            if mode[key]=='all':
                sheet = wb[key]
                for i in range(2,sheet.max_row+1):
                    row_data={}
                    row_data['case_id']=sheet.cell(i,1).value
                    row_data['url']=sheet.cell(i,2).value
                    #替换测试数据
                    # print(type(sheet.cell(2, 3).value))

                    if str(sheet.cell(i,3).value).find("${NoRegTel}")!=-1:
                        row_data['data'] = sheet.cell(i, 3).value.replace("${NoRegTel}", str(NoRegTel))
                        NoRegTel = NoRegTel + 1
                    else:
                        row_data['data']=DoRegx.do_regx(sheet.cell(i,3).value)
                    if sheet.cell(i,4).value !=None :
                        # sql语句的处理
                        row_data['check_sql'] = DoRegx.do_regx(sheet.cell(i,4).value)
                    else:
                        row_data['check_sql'] = None
                    row_data['title']=sheet.cell(i,5).value
                    row_data['http_method']=sheet.cell(i,6).value
                    row_data['expected']=sheet.cell(i,7).value
                    row_data['sheet_name']=key
                    test_data.append(row_data)
                    cls.updata(NoRegTel,filename,'init')

            else:
                sheet = wb[key]
                for case_id in mode[key]:
                    row_data = {}
                    row_data['case_id'] = sheet.cell(case_id+1,1).value
                    row_data['url'] = sheet.cell(case_id+1,2).value
                    # 替换测试数据
                    if str(sheet.cell(case_id+1,3).value).find("${tel}") != -1:
                        row_data['data'] = sheet.cell(case_id+1, 3).value.replace("${tel}", str(NoRegTel))
                        NoRegTel = NoRegTel + 1
                    else:
                        row_data['data']=DoRegx.do_regx(sheet.cell(case_id+1,3).value)

                    if sheet.cell(i,4).value !=None :
                        # sql语句的处理
                        row_data['check_sql'] = DoRegx.do_regx(sheet.cell(case_id+1,4).value)
                    else:
                        row_data['check_sql'] = None

                    row_data['title'] = sheet.cell(case_id+1,5).value
                    row_data['http_method'] = sheet.cell(case_id+1,6).value
                    row_data['expected'] = sheet.cell(case_id+1,7).value
                    row_data['sheetname']=key
                    test_data.append(row_data)
                    cls.updata(NoRegTel,filename,'init')

        return test_data

    @staticmethod
    def write_back(filename,sheetname,row,col,result):
        wb=load_workbook(filename)
        sheet = wb[sheetname]
        sheet.cell(row,col).value= result
        wb.save(filename)
        # wb.close()

    @classmethod
    def updata(cls,tel,filename,sheetname):
        wb = load_workbook(filename)
        sheet = wb[sheetname]
        sheet.cell(2,1).value = tel
        wb.save(filename)


if __name__ == '__main__':

    # wb = load_workbook(test_case_path)
    # sheet = wb["bussiness"]
    # test_data = []
    # row_data = {}
    # # row_data['case_id'] = sheet.cell(2, 1).value
    # # row_data['url'] = sheet.cell(2, 2).value
    # print(type(sheet.cell(2, 3).value))
    # if sheet.cell(2, 3).value.find("${tel}") != -1:
    #     print("没报错")
    # test_data.append(row_data)
    # print(test_data)

    test_data=DoExcel().get_data(test_case_path)
    print(test_data)

    # o=["hello!","world"]
    # for i in o:
    #     if i.find("h")!= -1:
    #         print('h')
    #     if i.find("e")!= -1:
    #         print("e")
    #     if i.find("!")!= -1:
    #         print("!")
    #     if i.find("h")==-1 and i.find("e")==-1 and i.find("!")==-1 :
    #         print("都没有")
